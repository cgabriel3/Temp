import argparse
import configparser
import logging
import os
from datetime import datetime, timedelta

from openpyxl import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import base64

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tapd import Tapd
from phabricator import Phabricator

tapd = None
phabricator = None

directory = os.path.dirname(__file__)
report_directory = None
log_directory = None

TAPD_TASK_DATA = []

INDO_DEVELOPER_USERNAME_LIST = [
  "aflah.nadhif",
  "andrey.martin",
  "barry.juans",
  "christian.gabriel.isjwara",
  "danny.restu",
  "dante.clericuzio",
  "deddy.zainal",
  "dewy.yuliana",
  "ega.frandika",
  "hafiz.siregar",
  "himawan.saputra.utama",
  "i.gede.agung",
  "james.surya.seputro",
  "kenneth.k.chang",
  "m.ridho.saputra",
  "mario.claudius",
  "steven.agustinus",
  "yulia.dewi"
]

NOT_ENGINEER_USERNAME_LIST = [
  "james.surya.seputro",
  "deddy.zainal",
  "yulia.dewi",
  "danny.restu",
  "steven.agustinus"
]

TASK_LINK = "https://www.tapd.cn/59680598/prong/stories/view/"
LOCAL_BUG_LINK = "https://www.tapd.cn/59680598/bugtrace/bugs/view/"
HQ_BUG_LINK = "https://www.tapd.cn/53182677/bugtrace/bugs/view/"


def generate_developer_name_empty_count_map():
  developer_username_to_count_map = {}
  for developer in INDO_DEVELOPER_USERNAME_LIST:
    if developer_username_to_count_map.get(developer) is None:
      developer_username_to_count_map[developer] = 0
  return developer_username_to_count_map


def transform_time(story_time_string):
  story_time_format = '%Y-%m-%d %H:%M:%S'
  story_modified_time = datetime.strptime(story_time_string, story_time_format)
  return story_modified_time


# Filter Lists
def filter_task_week(task):
  task_modified_time = transform_time(task['Task']['modified'])
  return task_modified_time > get_start_of_week_time()


def filter_task_semester(task):
  task_created_time = transform_time(task['Task']['created'])
  return task_created_time > get_start_of_semester_time()


def filter_bug(bug):
  bug_created_time = transform_time(bug["Bug"]["created"])
  return bug_created_time > get_start_of_semester_time()


# Hit Tapd API
def get_tapd_task(task_filter=None):
  global TAPD_TASK_DATA

  if len(TAPD_TASK_DATA) == 0:
    TAPD_TASK_DATA = tapd.get_all_task()

  if task_filter is None:
    return TAPD_TASK_DATA

  return list(filter(lambda x: task_filter(x), TAPD_TASK_DATA))


def get_tapd_bug(tapd_bug_data):
  tapd_filtered_testing_bug_data = filter(lambda x: filter_bug(x), tapd_bug_data)
  return list(tapd_filtered_testing_bug_data)


# Generate Developer to Count Data
def generate_bug_data(tapd_bug_list):
  # Initialize count to name list
  developer_username_to_count_map = {}
  for developer in INDO_DEVELOPER_USERNAME_LIST:
    if developer_username_to_count_map.get(developer) is None:
      developer_username_to_count_map[developer] = {
        "total_count": 0,
        "this_week": 0,
        "bug_ids": ""
      }

  start_of_week = get_start_of_week_time()
  for bug in tapd_bug_list:
    developers = generate_developer_list(bug["Bug"]["de"])
    if developers is None:
      continue

    if len(developers) > 0:
      for developer in developers:
        if developer in INDO_DEVELOPER_USERNAME_LIST:
          developer_username_to_count_map[developer]["total_count"] = developer_username_to_count_map[developer]["total_count"] + 1

          if transform_time(bug["Bug"]["created"]) > start_of_week:
            developer_username_to_count_map[developer]["this_week"] = developer_username_to_count_map[developer]["this_week"] + 1

          bug_ids = developer_username_to_count_map[developer]["bug_ids"]
          if bug_ids != "":
            bug_ids += "\n"

          developer_username_to_count_map[developer]["bug_ids"] = bug_ids + LOCAL_BUG_LINK + bug["Bug"]["id"]

  data = []
  for key, value in developer_username_to_count_map.items():
    developer_data = [key, value["total_count"], value["this_week"], value["bug_ids"]]
    data.append(developer_data)
  return data


def generate_production_bug_data(tapd_bug_list):
  # Initialize count to name list
  developer_username_to_count_map = {}
  for developer in INDO_DEVELOPER_USERNAME_LIST:
    if developer_username_to_count_map.get(developer) is None:
      developer_username_to_count_map[developer] = {
        "total_count": 0,
        "bug_ids": ""
      }

  for bug in tapd_bug_list:
    if bug["Bug"]["originphase"] != "已上线":
      continue
    developers = generate_developer_list(bug["Bug"]["de"])
    if developers is None:
      continue

    if len(developers) > 0:
      for developer in developers:
        if developer in INDO_DEVELOPER_USERNAME_LIST:
          developer_username_to_count_map[developer]["total_count"] = developer_username_to_count_map[developer]["total_count"] + 1
          bug_ids = developer_username_to_count_map[developer]["bug_ids"]
          if bug_ids != "":
            bug_ids += "\n"

          developer_username_to_count_map[developer]["bug_ids"] = bug_ids + HQ_BUG_LINK + bug["Bug"]["id"]

  data = []
  for key, value in developer_username_to_count_map.items():
    developer_data = [key, value["total_count"], value["bug_ids"]]
    data.append(developer_data)
  return data


def generate_developer_list(developers_str):
  if developers_str is None:
    return None
  return developers_str.split(";")


def generate_task_statistic_data():
  # Initialize count to name list
  developer_username_to_count_map = {}
  for developer in INDO_DEVELOPER_USERNAME_LIST:
    if developer_username_to_count_map.get(developer) is None:
      developer_username_to_count_map[developer] = {
        "not_start": 0,
        "in_process": 0,
        "completed": 0,
        "completed_this_week": 0,
        "total_task": 0,
        "total_days": 0,
        "task_delayed": 0,
        "delayed_task_ids": ""
      }

  logging.info("Get TAPD Task List")
  tapd_task_list = get_tapd_task(task_filter=filter_task_semester)
  start_of_week = get_start_of_week_time()
  for task in tapd_task_list:
    developers = generate_developer_list(task["Task"]["owner"])
    due_date = task["Task"]["due"]

    due_date_datetime = None
    if due_date is not None:
      due_date_datetime = datetime.strptime(due_date, "%Y-%m-%d")

    current_date = datetime.now()
    late = False

    if developers is None:
      continue

    for developer in developers:
      if developer in INDO_DEVELOPER_USERNAME_LIST:
        developer_username_to_count_map[developer]["total_task"] = developer_username_to_count_map[developer]["total_task"] + 1

        if task["Task"]["status"] == "done":
          developer_username_to_count_map[developer]["completed"] = developer_username_to_count_map[developer]["completed"] + 1
          completed_time = task["Task"]["completed"]

          if completed_time is None:
            continue

          completed_time_datetime = transform_time(completed_time)
          if due_date_datetime is not None:
            late = completed_time_datetime.date() > due_date_datetime.date()

          if completed_time_datetime > start_of_week:
            developer_username_to_count_map[developer]["completed_this_week"] = developer_username_to_count_map[developer]["completed_this_week"] + 1

          begin_date = task["Task"]["begin"]
          if begin_date is None:
            continue
          begin_time = datetime.strptime(begin_date, "%Y-%m-%d")
          delta = completed_time_datetime.date() - begin_time.date()
          total_days = delta.days
          developer_username_to_count_map[developer]["total_days"] = developer_username_to_count_map[developer]["total_days"] + total_days

        elif task["Task"]["status"] == "open":
          developer_username_to_count_map[developer]["not_start"] = developer_username_to_count_map[developer]["not_start"] + 1
          if due_date_datetime is not None:
            late = current_date.date() > due_date_datetime.date()

        else:
          developer_username_to_count_map[developer]["in_process"] = developer_username_to_count_map[developer]["in_process"] + 1
          if due_date_datetime is not None:
            late = current_date.date() > due_date_datetime.date()

        if late:
          developer_username_to_count_map[developer]["task_delayed"] = developer_username_to_count_map[developer]["task_delayed"] + 1
          delayed_tasks = developer_username_to_count_map[developer]["delayed_task_ids"]
          if delayed_tasks != "":
            delayed_tasks += "\n"
          developer_username_to_count_map[developer]["delayed_task_ids"] = delayed_tasks + TASK_LINK + task["Task"]["id"]

  data = []
  for key, value in developer_username_to_count_map.items():
    late_percentage = 0.0
    average_completion_time = 0
    delayed_task_amount = value["task_delayed"]
    task_completed = value.get("completed")
    total_task = value["total_task"]
    if delayed_task_amount > 0:
      late_percentage = delayed_task_amount / total_task

    if value["total_days"] != 0 and value["completed"] != 0:
      average_completion_time = value["total_days"] / task_completed
    developer_data = [
      key,  # Developer Name
      value.get("not_start"),  # Total Task Not Started
      value.get("in_process"),  # Total Task In Progress
      value.get("completed_this_week"),  # Total Task Completed This Week
      task_completed,  # Total Task Completed
      total_task,  # Total Task
      average_completion_time,  # Average Completion Time of completed Tasks
      delayed_task_amount,  # Delayed Task Amount
      late_percentage,  # Percentage of Late Tasks from total tasks
      value.get("delayed_task_ids")  # Delayed task IDs
    ]
    data.append(developer_data)
  return data


def generate_code_changes_statistics_data():
  developer_username_to_count_map = {}
  engineers = []
  for developer in INDO_DEVELOPER_USERNAME_LIST:
    if developer not in NOT_ENGINEER_USERNAME_LIST and developer_username_to_count_map.get(developer) is None:
      developer_username_to_count_map[developer] = {
        "lines_added": 0,
        "lines_removed": 0
      }
      engineers.append(developer)

  for engineer in engineers:
    developer_username_to_count_map = get_engineer_code_changes(engineer, developer_username_to_count_map)

  data = []
  for engineer in engineers:
    lines_added = developer_username_to_count_map[engineer]["lines_added"]
    lines_removed = developer_username_to_count_map[engineer]["lines_removed"]
    total_lines = lines_added + lines_removed
    delta_lines = lines_added - lines_removed
    developer_data = [engineer, lines_added, lines_removed, total_lines, delta_lines]
    data.append(developer_data)
  return data


def get_engineer_code_changes(engineer, developer_username_to_count_map):
  diff_list = phabricator.get_diff_list(engineer)
  if diff_list is None:
    logging.info(f'No Diffs for {engineer}')
    return developer_username_to_count_map

  start_of_week_time = get_start_of_week_time()
  start_of_week_timestamp = start_of_week_time.timestamp()

  logging.info(f'Get latest diffs for {engineer}')
  for diff in diff_list:
    if float(diff["dateCreated"]) < start_of_week_timestamp:
      logging.info(f'Finish getting diffs for {engineer}')
      return developer_username_to_count_map
    developer_username_to_count_map[engineer]["lines_added"] = developer_username_to_count_map[engineer]["lines_added"] + diff["properties"]["lines.added"]
    developer_username_to_count_map[engineer]["lines_removed"] = developer_username_to_count_map[engineer]["lines_removed"] + diff["properties"]["lines.removed"]

  return developer_username_to_count_map


def generate_testing_bug_sheet(wb):
  logging.info(f'Start generate testing bug sheet')
  title = "Testing Bug"
  headers = ["Developer Name", "Total Amount of Testing Bug", "Amount of Testing Bug created This Week", "Bug Ids"]

  logging.info("Get Testing Bug from Indonesia Local TAPD Project")
  tapd_bug_data = get_tapd_bug(tapd.get_indonesia_testing_bug())
  data = generate_bug_data(tapd_bug_data)
  return generate_worksheet(wb, title, headers, data, first_sheet_flag=True)


def generate_production_bug_sheet(wb):
  logging.info(f'Start generate production bug sheet')
  title = "Production Bug"
  headers = ["Developer Name", "Amount of Production Bug", "Bug Ids"]

  logging.info("Get Production Bug from Indonesia Project(HQ) TAPD Project")
  tapd_bug_data = get_tapd_bug(tapd.get_indonesia_production_bug())
  data = generate_production_bug_data(tapd_bug_data)
  return generate_worksheet(wb, title, headers, data)


def generate_task_statistic_sheet(wb):
  logging.info(f'Start generate task statistic sheet')
  title = "Task Statistics"
  headers = [
    "Developer Name",
    "Total Task Not Started",
    "Total Task In Progress",
    "Total Task Completed This Week",
    "Total Task Completed This Semester",
    "Total Tasks",
    "Average Time to Complete Tasks (in Days)",
    "Tasks Delayed",
    "Task Delay Percentage",
    "Task Delay Ids"
  ]
  data = generate_task_statistic_data()
  return generate_worksheet(wb, title, headers, data)


def generate_code_changes_statistics_sheet(wb):
  logging.info(f'Start generate code changes statistics sheet')
  title = "Delta Lines of Code Changes"
  headers = ["Developer Name", "Lines of Code Added", "Lines of Code Removed", "Total Lines of Code Changed", "Delta Lines of Code"]
  data = generate_code_changes_statistics_data()
  return generate_worksheet(wb, title, headers, data)


def generate_worksheet(wb, title, header, data, first_sheet_flag=False):
  ws = wb.active
  if not first_sheet_flag:
    ws = wb.create_sheet()
  ws.title = title

  # Font Styling
  header_font = Font(name='Arial', bold=True, color='FFFFFF')  # White text

  # Fill Color
  header_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')  # Blue background

  # Alignment
  centered_alignment = Alignment(horizontal='center', vertical='center')

  # Border Styling
  thin_border = Border(left=Side(style='thin'),
                       right=Side(style='thin'),
                       top=Side(style='thin'),
                       bottom=Side(style='thin'))

  # Apply styles to headers
  for col_num in range(1, len(header) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = centered_alignment
    cell.border = thin_border
    cell.value = header[col_num - 1]

  # Apply styles to data cells (optional)
  # Use a for loop with index to fill the cells manually
  for row_index, row in enumerate(data, start=2):  # start=2 to match Excel's row numbering
    for col_index, value in enumerate(row, start=1):  # start=1 for Excel's column numbering
      cell = ws.cell(row=row_index, column=col_index, value=value)
      cell.alignment = Alignment(wrap_text=True)

  # Adjust column widths to fit content
  for col_num, col_cells in enumerate(ws.columns, start=1):
    max_length = 0
    column_letter = get_column_letter(col_num)  # Get the column letter

    for cell in col_cells:
      if cell.value:
        max_length = max(max_length, len(str(cell.value)))  # Find the max length of cell content in the column

    # Set the column width based on the max length (a little padding can be added)
    adjusted_width = max_length + 2
    ws.column_dimensions[column_letter].width = adjusted_width

  logging.info(f'Finish generate {title}')
  return wb


def send_email_with_attachment(config, filename):
  # Get email config section
  section = 'email'

  # Define email parameters
  sender_email = config.get(section, "sender_email_address")
  receiver_email = config.get(section, "receiver_email_address")
  subject = f'Indonesia Developer Activity Report - {get_current_date_string()}'
  body = "This email contains the activity report of Fintopia Indonesia's Developers"
  attachment_path = filename  # Path to the Excel fil e
  smtp_server = config.get(section, "host")
  smtp_port = config.get(section, "port")
  encoded_password = config.get(section, "sender_email_password")
  sender_password = base64.b64decode(encoded_password).decode()

  # Create message container
  msg = MIMEMultipart()
  msg['From'] = sender_email
  msg['To'] = receiver_email
  msg['Subject'] = subject

  # Attach the email body
  msg.attach(MIMEText(body, 'plain'))

  # Attach the file
  with open(attachment_path, "rb") as attachment:
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())

  encoders.encode_base64(part)
  part.add_header(
    'Content-Disposition',
    f'attachment; filename= {os.path.basename(attachment_path)}',
  )

  msg.attach(part)

  # Send the email via SMTP server
  with smtplib.SMTP(smtp_server, smtp_port) as server:
    try:
      server.starttls()  # Encrypt the connection
      server.login(sender_email, sender_password)
      text = msg.as_string()
      server.sendmail(sender_email, receiver_email, text)
      logging.info("Email successfully sent to " + receiver_email)
    except Exception as e:
      logging.info("Fail to send email, Error: " + str(e))
    finally:
      server.quit()


def get_current_date_string():
  # Get the current date
  today = datetime.today()

  # Format the date as DD MM YYYY
  formatted_date = today.strftime("%d %m %Y")
  return formatted_date


def generate_filename():
  return f'{directory}{report_directory}activity_report_{get_current_date_string()}.xlsx'


def get_start_of_week_time():
  now = datetime.now()
  days_since_monday = now.weekday()
  monday = now - timedelta(days=days_since_monday)
  monday_start = datetime(monday.year, monday.month, monday.day)
  return monday_start


def get_start_of_semester_time():
  start_of_semester = datetime(2024, 7, 1, 0, 0, 0)
  return start_of_semester


def generate_workbook(filename):
  wb = Workbook()
  wb = generate_testing_bug_sheet(wb)
  wb = generate_production_bug_sheet(wb)
  wb = generate_task_statistic_sheet(wb)
  wb = generate_code_changes_statistics_sheet(wb)
  wb.save(filename)


def get_env(env):
  if not env:
    env = 'prod'
  return f'{directory}/config.{env}.ini'


def setup_logging():
  # Configure the logging module
  timestamp = datetime.now().strftime('%d-%m-%Y')
  sync_tasks_log_file = f'{directory}{log_directory}activity_report_log_{timestamp}.log'

  # Configure the logger
  logging.basicConfig(
    filename=sync_tasks_log_file,
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
  )


def setup_constants(config):
  global tapd
  tapd = Tapd(config)

  global phabricator
  phabricator = Phabricator(config)

  global report_directory
  report_directory = config.get("DEFAULT", "report_directory")

  global log_directory
  log_directory = config.get("DEFAULT", "log_directory")


def report_weekly_activity(config):
  logging.info("----Activity Report Script Start----")

  filename = generate_filename()
  generate_workbook(filename)
  send_email_with_attachment(config, filename)
  logging.info("----Activity Report Script Finish----")


def main():
  parser = argparse.ArgumentParser(description='Sync Task Script')
  parser.add_argument('--env', help='Environment')
  args = parser.parse_args()
  env = get_env(args.env)
  config = configparser.ConfigParser()
  config.read(env)

  setup_constants(config)
  setup_logging()
  logging.info("Setup configs")
  report_weekly_activity(config)


main()
